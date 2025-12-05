from flask import (
    Flask,
    render_template,
    request,
    redirect,
    url_for,
    flash,
    send_file,
)
from openpyxl import load_workbook, Workbook
from io import BytesIO
import math

app = Flask(__name__)
app.secret_key = "ganti_ini_dengan_string_random_yang_agak_panjang"

# Konstanta
ALLOWED_EXTENSIONS = {"xlsx"}
SHEET_NAME = "Worksheet"
ROWS_PER_PAGE = 50

# Penyimpanan sederhana data terakhir untuk keperluan preview & download
last_rekap_rows: list[dict] = []


# ==========================
#  UTILITAS UMUM
# ==========================
def allowed_file(filename: str) -> bool:
    """Cek apakah ekstensi file diizinkan (.xlsx)."""
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


# ==========================
#  LOGIC KONVERSI & PROSES DATA
# ==========================
def parse_late(value) -> int:
    """
    Konversi nilai 'Late In' menjadi menit (int).
    Mendukung format:
    - "HH:MM"
    - "10" (menit)
    - "10 menit", "T+5", dll (akan diambil digit angkanya saja).
    """
    if value is None:
        return 0

    s = str(value).strip()
    if not s:
        return 0

    # Format "HH:MM"
    if ":" in s:
        try:
            h, m = s.split(":")
            return int(h) * 60 + int(m)
        except Exception:
            return 0

    # Hanya digit
    if s.isdigit():
        return int(s)

    # Ambil digit dari teks campuran
    digits = "".join(c for c in s if c.isdigit())
    return int(digits) if digits else 0


def proses_absensi_worksheet(ws) -> list[dict]:
    """
    Transformasi raw worksheet menjadi list of dict berisi hanya baris yang terlambat.
    Menambahkan kolom 'Total_Keterlambatan' per karyawan.

    ws: worksheet dari openpyxl (sheet 'Worksheet')
    return: list[dict] (rekap_final)
    """
    # ========== Ambil header di baris pertama ==========
    header_row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    header_map = {
        str(h).strip(): idx
        for idx, h in enumerate(header_row)
        if h is not None
    }

    # Wajib ada kolom berikut:
    required_cols = ["Full Name", "Date", "Check In", "Check Out", "Late In", "Early Out"]
    for col in required_cols:
        if col not in header_map:
            raise ValueError(f"Kolom '{col}' tidak ditemukan di header Excel.")

    idx_fullname = header_map["Full Name"]
    idx_date = header_map["Date"]
    idx_checkin = header_map["Check In"]
    idx_checkout = header_map["Check Out"]
    idx_latein = header_map["Late In"]
    idx_earlyout = header_map["Early Out"]

    # ========== Baca semua baris, bersihkan nama, hitung telat ==========
    rows_late: list[dict] = []
    count_per_name: dict[str, int] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        full_name_raw = row[idx_fullname]

        # Bersihkan nama
        if full_name_raw is None:
            continue

        full_name = " ".join(str(full_name_raw).split()).strip()
        if not full_name:
            continue
        if full_name.lower() == "nan":
            continue

        date_val = row[idx_date]
        check_in = row[idx_checkin]
        check_out = row[idx_checkout]
        late_in = row[idx_latein]
        early_out = row[idx_earlyout]

        late_minutes = parse_late(late_in)
        is_late = late_minutes > 0

        # Hanya simpan baris yang terlambat
        if not is_late:
            continue

        data_row = {
            "Full Name": full_name,
            "Date": date_val,
            "Check In": check_in,
            "Check Out": check_out,
            "Late In": late_in,
            "Early Out": early_out,
        }
        rows_late.append(data_row)

        # Hitung total telat per orang
        count_per_name[full_name] = count_per_name.get(full_name, 0) + 1

    # Kalau tidak ada yang telat, kembalikan list kosong
    if not rows_late:
        return []

    # Tambahkan kolom Total_Keterlambatan ke setiap row
    for r in rows_late:
        r["Total_Keterlambatan"] = count_per_name.get(r["Full Name"], 0)

    # Sort berdasarkan nama, lalu tanggal
    def sort_key(r: dict):
        name_key = r["Full Name"].lower()
        date_val = r["Date"]
        return (name_key, date_val if date_val is not None else "")

    rows_late.sort(key=sort_key)
    return rows_late


def buat_workbook_rekap(rekap_rows: list[dict]) -> Workbook:
    """
    Membuat workbook Excel dari list rekap_rows.

    rekap_rows: list[dict] dengan key:
      Full Name, Total_Keterlambatan, Date, Check In, Check Out, Late In, Early Out
    """
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Rekap Telat"

    headers = [
        "Full Name",
        "Total_Keterlambatan",
        "Date",
        "Check In",
        "Check Out",
        "Late In",
        "Early Out",
    ]
    ws_out.append(headers)

    for r in rekap_rows:
        ws_out.append(
            [
                r.get("Full Name", ""),
                r.get("Total_Keterlambatan", 0),
                r.get("Date", ""),
                r.get("Check In", ""),
                r.get("Check Out", ""),
                r.get("Late In", ""),
                r.get("Early Out", ""),
            ]
        )

    return wb_out


# ==========================
#  RINGKASAN + FILTER + PAGINATION
# ==========================
def hitung_ringkasan(rekap_rows: list[dict]) -> dict:
    """
    Hitung ringkasan:
    - total kejadian telat
    - total karyawan yang pernah telat
    - rata-rata telat per karyawan
    - top 5 karyawan paling sering telat
    """
    total_cases = len(rekap_rows)

    if not rekap_rows:
        return {
            "total_cases": 0,
            "total_employees": 0,
            "avg_per_employee": 0.0,
            "top5": [],
        }

    # Ambil total_keterlambatan per nama (satu kali per orang)
    counts: dict[str, int] = {}
    for r in rekap_rows:
        name = r.get("Full Name", "")
        if not name:
            continue
        count = r.get("Total_Keterlambatan", 0)
        # Kalau ada duplikat nama, ambil nilai terbesar (harusnya sama semua)
        if name not in counts or count > counts[name]:
            counts[name] = count

    total_employees = len(counts)
    avg_per_employee = total_cases / total_employees if total_employees > 0 else 0.0

    top5_raw = sorted(counts.items(), key=lambda x: x[1], reverse=True)[:5]
    top5 = [{"name": n, "count": c} for n, c in top5_raw]

    return {
        "total_cases": total_cases,
        "total_employees": total_employees,
        "avg_per_employee": avg_per_employee,
        "top5": top5,
    }


def paginate_rows(rows: list[dict], page: int, per_page: int = ROWS_PER_PAGE):
    """Bagi rows menjadi beberapa halaman, mengembalikan subset + info pagination."""
    total = len(rows)
    if total == 0:
        return [], {
            "page": 1,
            "total_pages": 1,
            "per_page": per_page,
            "total": 0,
            "start_index": 0,
            "end_index": 0,
        }

    total_pages = math.ceil(total / per_page)
    page = max(1, min(page, total_pages))

    start = (page - 1) * per_page
    end = start + per_page
    subset = rows[start:end]

    return subset, {
        "page": page,
        "total_pages": total_pages,
        "per_page": per_page,
        "total": total,
        "start_index": start + 1,
        "end_index": min(end, total),
    }


def filter_rows(rows: list[dict], query: str) -> list[dict]:
    """
    Filter sederhana berdasarkan nama karyawan (Full Name).
    query: string (boleh kosong / None).
    """
    if not query:
        return rows

    q = query.strip().lower()
    if not q:
        return rows

    filtered: list[dict] = []
    for r in rows:
        name = str(r.get("Full Name", "")).lower()
        if q in name:
            filtered.append(r)
    return filtered


# ==========================
#  ROUTES WEB
# ==========================
@app.route("/", methods=["GET"])
def index():
    """Halaman upload file absensi."""
    return render_template("index.html")


@app.route("/process", methods=["POST"])
def process():
    """Menerima upload file, memproses, lalu redirect ke halaman preview."""
    global last_rekap_rows

    if "file" not in request.files:
        flash("File tidak ditemukan di request.")
        return redirect(url_for("index"))

    file = request.files["file"]

    if file.filename == "":
        flash("Silakan pilih file absensi (.xlsx) terlebih dahulu.")
        return redirect(url_for("index"))

    if not allowed_file(file.filename):
        flash("Format file harus .xlsx")
        return redirect(url_for("index"))

    try:
        wb = load_workbook(file, data_only=True)
    except Exception as e:
        flash(f"Gagal membaca file Excel: {e}")
        return redirect(url_for("index"))

    if SHEET_NAME not in wb.sheetnames:
        flash(f'Sheet "{SHEET_NAME}" tidak ditemukan di file Excel.')
        return redirect(url_for("index"))

    ws = wb[SHEET_NAME]

    try:
        rekap_rows = proses_absensi_worksheet(ws)
    except Exception as e:
        flash(f"Terjadi error saat memproses data: {e}")
        return redirect(url_for("index"))

    # Simpan data yang sudah diproses di memori
    last_rekap_rows = rekap_rows

    # Setelah proses, arahkan ke halaman preview page 1
    return redirect(url_for("preview", page=1))


@app.route("/preview", methods=["GET"])
def preview():
    """Halaman preview tabel rekap (dengan filter dan pagination)."""
    global last_rekap_rows

    if not last_rekap_rows:
        flash("Belum ada data yang diproses.")
        return redirect(url_for("index"))

    # Ambil query & page dari URL
    q = request.args.get("q", default="", type=str)
    page = request.args.get("page", default=1, type=int)

    # Filter berdasarkan nama (jika ada)
    filtered_rows = filter_rows(last_rekap_rows, q)

    # Pagination berdasarkan hasil filter
    paged_rows, page_info = paginate_rows(filtered_rows, page, ROWS_PER_PAGE)

    # Ringkasan berdasarkan data yang sudah difilter
    summary = hitung_ringkasan(filtered_rows)

    return render_template(
        "preview.html",
        rows=paged_rows,
        summary=summary,
        page_info=page_info,
        q=q,  # kirim query ke template supaya input-nya bisa diisi ulang
    )


@app.route("/download", methods=["GET"])
def download():
    """Download seluruh rekap (tanpa filter & pagination) sebagai Excel."""
    global last_rekap_rows

    if not last_rekap_rows:
        flash("Belum ada data yang diproses.")
        return redirect(url_for("index"))

    wb_out = buat_workbook_rekap(last_rekap_rows)

    output = BytesIO()
    wb_out.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="rekap_semua_telat_clean.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    app.run(debug=False)
